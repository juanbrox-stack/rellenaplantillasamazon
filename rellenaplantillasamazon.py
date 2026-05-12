import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import re

st.set_page_config(page_title="Amazon Pro Filler", layout="wide")

st.title("🚀 Amazon Template Automator")
st.markdown("Configurado para detectar nombres técnicos en **Fila 4** y escribir desde **Fila 7**.")

# --- VALORES FIJOS (No cambian) ---
VALORES_FIJOS = {
    "brand#1.value": "Cecotec", "external_product_id#1.type": "EAN", 
    "package_level#1.value": "Unit", "is_trade_item_orderable_unit#1.value": "No",
    "manufacturer#1.value": "Cecotec", "number_of_items#1.value": 1,
    "is_oem_authorized#1.value": "Yes", "wattage#1.unit": "Watts",
    "item_depth_width_height#1.depth.unit": "Centimeters",
    "item_depth_width_height#1.height.unit": "Centimeters",
    "item_depth_width_height#1.width.unit": "Centimeters",
    "item_dimensions#1.length.unit": "Centimeters",
    "item_dimensions#1.width.unit": "Centimeters",
    "item_dimensions#1.height.unit": "Centimeters",
    "item_package_dimensions#1.length.unit": "Centimeters",
    "item_package_dimensions#1.width.unit": "Centimeters",
    "item_package_dimensions#1.height.unit": "Centimeters",
    "item_package_weight#1.unit": "Kilograms", "rtip_items_per_inner_pack#1.value": 1,
    "country_of_origin#1.value": "Spain", "warranty_description#1.value": "2 ans du garantie",
    "supplier_declared_dg_hz_regulation#1.value": "Not Applicable",
    "item_weight#1.unit": "Kilograms", "eu_spare_part_availability_duration#1.value": 10,
    "eu_spare_part_availability_duration#1.unit": "Years",
    "dsa_responsible_party_address#1.value": "https://cecotec.es/",
    "compliance_media#1.content_type": "User Manual",
    "compliance_media#1.content_language": "fr_FR", "gpsr_safety_attestation#1.value": "Yes",
    "gpsr_manufacturer_reference#1.gpsr_manufacturer_email_address": "https://cecotec.es/"
}

# --- MAPEO VARIABLE (Configuración inicial) ---
MAPEO_INICIAL = {
    "vendor_sku#1.value": "SKU", "external_product_id#1.value": "EAN",
    "merchant_suggested_asin#1.value": "ASIN", "model_number#1.value": "Nombre Producto / Modelo",
    "bullet_point#1.value": "Bulletpoint 1 (FR)", "bullet_point#2.value": "Bulletpoint 2 (FR)",
    "bullet_point#3.value": "Bulletpoint 3 (FR)", "bullet_point#4.value": "Bulletpoint 4 (FR)",
    "bullet_point#5.value": "Bulletpoint 5 (FR)", "item_type_name#1.value": "Subfamilia (FR)",
    "rtip_product_description#1.value": "Descripción larga del producto (FR)",
    "color#1.value": "color", "part_number#1.value": "SKU",
    "oem_equivalent_part_number#1.value": "SKU", "wattage#1.value": "Potencia (W)",
    "included_components#1.value": "Contenido de la caja (FR)",
    "item_depth_width_height#1.depth.value": "Product depth (cm)",
    "item_depth_width_height#1.height.value": "Product height (cm)",
    "item_depth_width_height#1.width.value": "Product width (cm)",
    "website_shipping_weight#1.value": "Peso Caja Color (kg)",
    "item_dimensions#1.length.value": "Product depth (cm)",
    "item_dimensions#1.width.value": "Product width (cm)",
    "item_dimensions#1.height.value": "Product height (cm)",
    "item_package_dimensions#1.length.value": "Largo Caja Color cm",
    "item_package_dimensions#1.width.value": "Ancho Caja Color cm",
    "item_package_dimensions#1.height.value": "Alto Caja Color cm",
    "item_package_weight#1.value": "Peso Caja Color (kg)",
    "item_weight#1.value": "Product weight (Kg)",
    "compliance_media#1.source_location": "Manual de instrucciones (Comercial)"
}

# --- BARRA LATERAL PARA MAPEO PERSONALIZADO ---
st.sidebar.header("⚙️ Ajustes de Mapeo")
st.sidebar.info("Asigna las columnas de tu PIM a los campos de Amazon.")

final_mapping = {}
with st.sidebar.expander("📝 Editar Correspondencias", expanded=False):
    for amz_key, pim_default in MAPEO_INICIAL.items():
        # Permitimos al usuario cambiar el nombre de la columna del PIM
        user_input = st.text_input(f"Amazon: {amz_key}", value=pim_default, key=amz_key)
        final_mapping[amz_key] = user_input

# --- FUNCIONES DE LIMPIEZA ---
def clean_header(text):
    if not text: return ""
    text = str(text).lower().strip()
    text = re.sub(r'#\d+', '', text) # Quita #1, #2...
    text = re.sub(r'\.value|\.unit|\.type', '', text) # Quita extensiones
    return re.sub(r'[^a-z0-9]', '', text) # Deja solo letras y números

# --- CARGA DE ARCHIVOS ---
col1, col2 = st.columns(2)
with col1:
    pim_file = st.file_uploader("📂 1. Subir Datos PIM (Origen)", type=["xlsx", "csv"])
with col2:
    amz_template = st.file_uploader("📂 2. Subir Plantilla Amazon (Destino)", type=["xlsx"])

if st.button("🚀 Ejecutar y Generar Fichero"):
    if pim_file and amz_template:
        try:
            # 1. Leer archivos
            df_pim = pd.read_excel(pim_file) if pim_file.name.endswith('.xlsx') else pd.read_csv(pim_file)
            wb = load_workbook(amz_template)
            ws = wb["Template"] if "Template" in wb.sheetnames else wb.worksheets[1]
            
            # 2. Mapear Fila 4 (Amazon)
            amz_indices_raw = {}
            amz_indices_clean = {}
            for i, cell in enumerate(ws[4], 1): # Mirar Fila 4
                if cell.value:
                    val = str(cell.value).strip()
                    amz_indices_raw[val] = i
                    amz_indices_clean[clean_header(val)] = i
            
            # 3. Procesar Filas (Desde Fila 7)
            log_mapeo = []
            for i, row_pim in df_pim.iterrows():
                target_row = i + 7
                
                # A. Valores del PIM (Mapeados en la barra lateral)
                for amz_key, pim_col in final_mapping.items():
                    # Buscar columna en Amazon (exacto o limpio)
                    idx = amz_indices_raw.get(amz_key) or amz_indices_clean.get(clean_header(amz_key))
                    
                    if idx and pim_col in df_pim.columns:
                        ws.cell(row=target_row, column=idx).value = row_pim[pim_col]
                        if i == 0: log_mapeo.append(f"✅ {amz_key} asignado a columna {idx}")

                # B. Valores Fijos
                for amz_key, fixed_val in VALORES_FIJOS.items():
                    idx = amz_indices_raw.get(amz_key) or amz_indices_clean.get(clean_header(amz_key))
                    if idx:
                        ws.cell(row=target_row, column=idx).value = fixed_val

            # 4. Mostrar resultados y descargar
            st.success(f"✅ Procesados {len(df_pim)} productos.")
            with st.expander("Ver detalle de emparejamiento"):
                st.write(log_mapeo)
                
            output = BytesIO()
            wb.save(output)
            st.download_button("📥 Descargar Plantilla Rellena", output.getvalue(), "Amazon_Carga_Masiva.xlsx")

        except Exception as e:
            st.error(f"Error: {e}")
    else:
        st.warning("Debes subir ambos archivos.")